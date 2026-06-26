"""文档摄入管道：解析 -> 切片 -> 向量化入库。

原则：全部用 LangChain 原生组件（Loaders + Splitter + VectorStore），
不自己写解析逻辑。解析/切片在可超时并硬终止的独立进程中执行。
"""
from __future__ import annotations

import asyncio
import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from langchain_community.document_loaders import (
    Docx2txtLoader,
    PyPDFLoader,
    TextLoader,
)
from langchain_community.vectorstores.utils import filter_complex_metadata
from langchain_core.document_loaders import BaseLoader
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from loguru import logger
from openpyxl import load_workbook

from app.config import settings
from app.core.process_pool import run_in_parser_process
from app.core.vectorstore import close_vectorstore, get_vectorstore, vectorstore_lock


@dataclass
class IngestResult:
    """摄入结果。"""

    success: bool
    chunk_count: int = 0
    error_msg: str | None = None
    trusted_path: str | None = None


class ParseLimitError(ValueError):
    """解析结果超过配置边界。"""


PARSE_LIMIT_MESSAGE = "解析文本超过字符数上限"


# 切分器：中英文混排友好的分隔符优先级
_SPLITTER = RecursiveCharacterTextSplitter(
    chunk_size=800,
    chunk_overlap=100,
    separators=["\n\n", "\n", "。", "!", "?", ",", " ", ""],
)


class OpenpyxlExcelLoader(BaseLoader):
    """轻量 XLSX Loader，输出标准 LangChain Document。

    每个工作表对应一个 Document，单元格按行转成制表符分隔文本。
    这避免引入 unstructured 的 OCR/Spacy 等与 Excel 无关的重依赖。
    """

    def __init__(self, file_path: str) -> None:
        self.file_path = file_path

    def lazy_load(self) -> Iterator[Document]:
        workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                lines: list[str] = []
                for row in worksheet.iter_rows(values_only=True):
                    values = ["" if value is None else str(value) for value in row]
                    if any(values):
                        lines.append("\t".join(values).rstrip())
                if lines:
                    yield Document(
                        page_content="\n".join(lines),
                        metadata={
                            "source": self.file_path,
                            "sheet_name": worksheet.title,
                        },
                    )
        finally:
            workbook.close()


def _select_loader(path: Path):
    """按扩展名选择 LangChain 原生 Loader。"""
    ext = path.suffix.lower()
    if ext == ".pdf":
        return PyPDFLoader(str(path))
    if ext == ".docx":
        return Docx2txtLoader(str(path))
    if ext == ".xlsx":
        return OpenpyxlExcelLoader(str(path))
    if ext in {".txt", ".md"}:
        # 先尝试 UTF-8 BOM，再回退 GBK（Windows 简体中文默认编码）
        for enc in ("utf-8-sig", "gbk"):
            try:
                path.read_text(encoding=enc)
                return TextLoader(str(path), encoding=enc)
            except UnicodeDecodeError:
                continue
        return TextLoader(str(path), encoding="utf-8-sig")
    raise ValueError(f"不支持的文件类型：{ext}")


def _load_and_split(
    path: Path,
    doc_id: int,
    source: str,
    tenant_id: str,
    uploaded_by: str,
    max_parsed_chars: int,
) -> list[Document]:
    """同步：加载 + 切片 + metadata；仅在独立解析进程中调用。"""
    loader = _select_loader(path)
    raw_docs = loader.load()
    total_chars = sum(len(document.page_content) for document in raw_docs)
    if total_chars > max_parsed_chars:
        raise ParseLimitError(PARSE_LIMIT_MESSAGE)
    chunks = _SPLITTER.split_documents(raw_docs)
    for index, chunk in enumerate(chunks):
        content_hash = hashlib.sha256(chunk.page_content.encode("utf-8")).hexdigest()
        chunk_id = f"{tenant_id}:{doc_id}:{index}:{content_hash}"
        chunk.metadata.update(
            {
                "doc_id": doc_id,
                "chunk_id": chunk_id,
                "content_hash": content_hash,
                "source": source,
                "chunk_index": index,
                "tenant_id": tenant_id,
                "uploaded_by": uploaded_by,
            }
        )
    # 清掉 loader 可能带入的复杂/None metadata，避免 Chroma 写入报错
    return filter_complex_metadata(chunks)


async def ingest_document(
    *,
    doc_id: int,
    file_path: str,
    source: str,
    tenant_id: str,
    uploaded_by: str,
) -> IngestResult:
    """摄入单个文档：解析、切片、异步写入向量库。

    Args:
        doc_id: 数据库中的文档主键，写入每个 chunk 的 metadata。
        file_path: 落盘文件路径。
        source: 展示用源文件名（写入 metadata.source）。
        tenant_id: 已验证 JWT 中的租户标识，用于向量检索隔离。
        uploaded_by: 已验证 JWT 中的用户标识，用于审计。

    Returns:
        IngestResult: 含成功标志、切片数、错误信息。
    """
    path = Path(file_path)
    logger.info("开始摄入 doc_id={} source={}", doc_id, source)
    try:
        chunks = await run_in_parser_process(
            _load_and_split,
            path,
            doc_id,
            source,
            tenant_id,
            uploaded_by,
            settings.max_parsed_chars,
            timeout=settings.parser_timeout_seconds,
        )
    except TimeoutError:
        logger.error("文档解析超时 doc_id={}", doc_id)
        return IngestResult(success=False, error_msg="文档解析超时")
    except ParseLimitError:
        logger.warning("文档超过解析限制 doc_id={}", doc_id)
        return IngestResult(success=False, error_msg=PARSE_LIMIT_MESSAGE)
    except Exception:  # noqa: BLE001
        logger.exception("文档解析失败 doc_id={}", doc_id)
        return IngestResult(success=False, error_msg="文档解析失败")

    if not chunks:
        return IngestResult(success=False, error_msg="未解析出任何文本内容")

    try:
        trusted_dir = settings.storage_dir / "documents" / tenant_id
        await asyncio.to_thread(trusted_dir.mkdir, parents=True, exist_ok=True)
        trusted_path = trusted_dir / path.name
        if path.resolve() != trusted_path.resolve():
            await asyncio.to_thread(path.replace, trusted_path)
    except OSError:
        logger.exception("隔离文件归档失败 doc_id={}", doc_id)
        return IngestResult(success=False, error_msg="文件归档失败")

    try:
        with vectorstore_lock():
            try:
                vectorstore = get_vectorstore()
                await vectorstore.aadd_documents(
                    chunks,
                    ids=[str(chunk.metadata["chunk_id"]) for chunk in chunks],
                )
            finally:
                close_vectorstore()

        logger.info("摄入完成 doc_id={} chunks={}", doc_id, len(chunks))
        return IngestResult(
            success=True,
            chunk_count=len(chunks),
            trusted_path=str(trusted_path),
        )
    except Exception:  # noqa: BLE001
        logger.exception("向量索引失败 doc_id={}", doc_id)
        return IngestResult(
            success=False,
            error_msg="文档索引失败",
            trusted_path=str(trusted_path),
        )
