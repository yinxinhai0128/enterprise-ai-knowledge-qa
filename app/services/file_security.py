"""上传文件的名称、格式、展开资源和恶意软件扫描边界。"""
from __future__ import annotations

import codecs
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Protocol

from openpyxl import load_workbook
from pypdf import PdfReader

from app.config import settings

ALLOWED_EXTENSIONS = frozenset({".pdf", ".docx", ".xlsx", ".txt"})
_ALLOWED_MIME_TYPES = {
    ".pdf": frozenset({"application/pdf", "application/octet-stream"}),
    ".docx": frozenset(
        {
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/zip",
            "application/octet-stream",
        }
    ),
    ".xlsx": frozenset(
        {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/zip",
            "application/octet-stream",
        }
    ),
    ".txt": frozenset({"text/plain", "application/octet-stream"}),
}
_WINDOWS_RESERVED = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}


class FileValidationError(Exception):
    """可安全返回客户端的文件拒绝原因。"""

    def __init__(self, message: str, status_code: int = 400) -> None:
        # 将 status_code 放入 args，保证跨 ProcessPool pickle 后仍保留状态码。
        super().__init__(message, status_code)
        self.safe_message = message
        self.status_code = status_code


@dataclass(frozen=True, slots=True)
class FileValidationLimits:
    max_archive_entries: int
    max_archive_uncompressed_bytes: int
    max_archive_compression_ratio: float
    max_pdf_pages: int
    max_xlsx_sheets: int
    max_xlsx_cells: int
    max_parsed_chars: int
    read_chunk_bytes: int


def current_file_validation_limits() -> FileValidationLimits:
    """在 API 进程读取当前配置，并显式传给隔离 Worker。"""
    return FileValidationLimits(
        max_archive_entries=settings.max_archive_entries,
        max_archive_uncompressed_bytes=settings.max_archive_uncompressed_bytes,
        max_archive_compression_ratio=settings.max_archive_compression_ratio,
        max_pdf_pages=settings.max_pdf_pages,
        max_xlsx_sheets=settings.max_xlsx_sheets,
        max_xlsx_cells=settings.max_xlsx_cells,
        max_parsed_chars=settings.max_parsed_chars,
        read_chunk_bytes=settings.upload_chunk_bytes,
    )


def normalize_filename(raw_name: str | None) -> tuple[str, str]:
    """规范化文件名、移除客户端路径并执行长度/字符/扩展名校验。"""
    normalized = unicodedata.normalize("NFKC", raw_name or "").strip()
    name = PurePosixPath(normalized.replace("\\", "/")).name
    if not name or name in {".", ".."}:
        raise FileValidationError("文件名无效")
    if len(name) > settings.max_filename_chars:
        raise FileValidationError(
            f"文件名超过 {settings.max_filename_chars} 字符上限"
        )
    if any(ord(char) < 32 or char == "\x7f" for char in name):
        raise FileValidationError("文件名包含非法控制字符")
    if Path(name).stem.upper() in _WINDOWS_RESERVED:
        raise FileValidationError("文件名为系统保留名称")
    ext = Path(name).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise FileValidationError(
            f"不支持的文件类型：{ext or '未知'}，仅支持 pdf/docx/xlsx/txt"
        )
    return name, ext


def _validate_mime(ext: str, content_type: str | None) -> None:
    mime = (content_type or "application/octet-stream").split(";", 1)[0].lower()
    if mime not in _ALLOWED_MIME_TYPES[ext]:
        raise FileValidationError("文件 MIME 类型与扩展名不匹配")


def _validate_zip(path: Path, ext: str, limits: FileValidationLimits) -> None:
    if not zipfile.is_zipfile(path):
        raise FileValidationError("文件内容与扩展名不匹配")
    required_member = "word/document.xml" if ext == ".docx" else "xl/workbook.xml"
    try:
        with zipfile.ZipFile(path) as archive:
            entries = archive.infolist()
            if len(entries) > limits.max_archive_entries:
                raise FileValidationError("压缩包文件条目过多")
            names = {entry.filename for entry in entries}
            if "[Content_Types].xml" not in names or required_member not in names:
                raise FileValidationError("文件内容与 Office 扩展名不匹配")
            total_uncompressed = 0
            for entry in entries:
                member = PurePosixPath(entry.filename.replace("\\", "/"))
                if member.is_absolute() or ".." in member.parts:
                    raise FileValidationError("压缩包包含非法路径")
                if entry.flag_bits & 0x1:
                    raise FileValidationError("不支持加密的 Office 文件")
                total_uncompressed += entry.file_size
                if total_uncompressed > limits.max_archive_uncompressed_bytes:
                    raise FileValidationError("压缩包展开后超过大小上限", 413)
                if entry.file_size:
                    ratio = entry.file_size / max(entry.compress_size, 1)
                    if ratio > limits.max_archive_compression_ratio:
                        raise FileValidationError("压缩包压缩比超过安全上限")
    except FileValidationError:
        raise
    except (OSError, zipfile.BadZipFile, RuntimeError) as exc:
        raise FileValidationError("Office 文件结构无效") from exc


def _validate_pdf(path: Path, limits: FileValidationLimits) -> None:
    try:
        with path.open("rb") as stream:
            if not stream.read(5).startswith(b"%PDF-"):
                raise FileValidationError("文件内容与 PDF 扩展名不匹配")
            stream.seek(0)
            reader = PdfReader(stream, strict=False)
            if reader.is_encrypted:
                raise FileValidationError("不支持加密 PDF")
            if len(reader.pages) > limits.max_pdf_pages:
                raise FileValidationError(
                    f"PDF 超过 {limits.max_pdf_pages} 页上限",
                    413,
                )
    except FileValidationError:
        raise
    except Exception as exc:  # pypdf 的异常类型跨版本不稳定
        raise FileValidationError("PDF 文件结构无效") from exc


def _validate_txt(path: Path, limits: FileValidationLimits) -> None:
    decoder = codecs.getincrementaldecoder("utf-8")("strict")
    total_chars = 0
    try:
        with path.open("rb") as stream:
            while chunk := stream.read(limits.read_chunk_bytes):
                if b"\x00" in chunk:
                    raise FileValidationError("TXT 文件包含二进制内容")
                total_chars += len(decoder.decode(chunk))
                if total_chars > limits.max_parsed_chars:
                    raise FileValidationError("文本内容超过字符数上限", 413)
            total_chars += len(decoder.decode(b"", final=True))
    except UnicodeDecodeError as exc:
        raise FileValidationError("TXT 文件必须使用 UTF-8 编码") from exc


def _validate_xlsx_shape(path: Path, limits: FileValidationLimits) -> None:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if len(workbook.worksheets) > limits.max_xlsx_sheets:
                raise FileValidationError(
                    f"XLSX 超过 {limits.max_xlsx_sheets} 个工作表上限",
                    413,
                )
            total_cells = 0
            total_chars = 0
            for worksheet in workbook.worksheets:
                total_cells += worksheet.max_row * worksheet.max_column
                if total_cells > limits.max_xlsx_cells:
                    raise FileValidationError("XLSX 单元格数量超过上限", 413)
                for row in worksheet.iter_rows(values_only=True):
                    total_chars += sum(len(str(value)) for value in row if value is not None)
                    if total_chars > limits.max_parsed_chars:
                        raise FileValidationError("XLSX 文本内容超过字符数上限", 413)
        finally:
            workbook.close()
    except FileValidationError:
        raise
    except Exception as exc:  # openpyxl 的异常类型跨版本不稳定
        raise FileValidationError("XLSX 文件结构无效") from exc


def validate_uploaded_file(
    path: Path,
    ext: str,
    content_type: str | None,
    limits: FileValidationLimits,
) -> None:
    """同步执行格式与资源校验；调用方应在隔离进程中设置超时。"""
    _validate_mime(ext, content_type)
    if ext == ".pdf":
        _validate_pdf(path, limits)
    elif ext in {".docx", ".xlsx"}:
        _validate_zip(path, ext, limits)
        if ext == ".xlsx":
            _validate_xlsx_shape(path, limits)
    else:
        _validate_txt(path, limits)


@dataclass(frozen=True, slots=True)
class MalwareScanResult:
    clean: bool
    scanned: bool


class MalwareScanner(Protocol):
    async def scan(self, path: Path) -> MalwareScanResult:
        """扫描隔离区文件，不得移动或修改原文件。"""


class DisabledMalwareScanner:
    async def scan(self, path: Path) -> MalwareScanResult:
        return MalwareScanResult(clean=True, scanned=False)


_malware_scanner: MalwareScanner = DisabledMalwareScanner()


def configure_malware_scanner(scanner: MalwareScanner) -> None:
    """供部署集成 ClamAV/EDR 等扫描器。"""
    global _malware_scanner
    _malware_scanner = scanner


def reset_malware_scanner() -> None:
    """恢复默认扫描器；主要用于测试隔离。"""
    global _malware_scanner
    _malware_scanner = DisabledMalwareScanner()


async def scan_quarantined_file(path: Path) -> None:
    result = await _malware_scanner.scan(path)
    if not result.clean:
        raise FileValidationError("文件未通过恶意软件扫描")
    if settings.malware_scan_required and not result.scanned:
        raise FileValidationError("恶意软件扫描服务不可用", 503)
