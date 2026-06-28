"""文档摄入接口测试：上传成功、类型校验、落库。"""
from __future__ import annotations

import asyncio
import io
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape

import pytest
from openpyxl import Workbook
from pypdf import PdfWriter
from pypdf.generic import DecodedStreamObject, DictionaryObject, NameObject
from sqlalchemy import select

import app.core.database as database_module
from app.config import settings
from app.core.process_pool import get_parser_pool
from app.models.document import Document
from app.services.file_security import (
    MalwareScanResult,
    configure_malware_scanner,
    reset_malware_scanner,
)
from app.services.ingest import IngestResult, ingest_document


def _pdf_bytes(pages: int) -> bytes:
    output = io.BytesIO()
    writer = PdfWriter()
    for _ in range(pages):
        writer.add_blank_page(width=100, height=100)
    writer.write(output)
    return output.getvalue()


def _text_pdf_bytes(text: str) -> bytes:
    """生成带真实可提取文本流的最小 PDF，不依赖外部样本文件。"""
    output = io.BytesIO()
    writer = PdfWriter()
    page = writer.add_blank_page(width=300, height=100)
    font = DictionaryObject(
        {
            NameObject("/Type"): NameObject("/Font"),
            NameObject("/Subtype"): NameObject("/Type1"),
            NameObject("/BaseFont"): NameObject("/Helvetica"),
        }
    )
    resources = DictionaryObject(
        {
            NameObject("/Font"): DictionaryObject(
                {NameObject("/F1"): writer._add_object(font)}
            )
        }
    )
    stream = DecodedStreamObject()
    escaped = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    stream.set_data(f"BT /F1 12 Tf 20 50 Td ({escaped}) Tj ET".encode("ascii"))
    page[NameObject("/Resources")] = resources
    page[NameObject("/Contents")] = writer._add_object(stream)
    writer.write(output)
    return output.getvalue()


def _xlsx_bytes(*, cells: int = 1, sheets: int = 1, value: str = "x") -> bytes:
    output = io.BytesIO()
    workbook = Workbook()
    first = workbook.active
    first.title = "sheet-1"
    for index in range(cells):
        first.cell(row=1, column=index + 1, value=value)
    for index in range(1, sheets):
        workbook.create_sheet(f"sheet-{index + 1}")
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _oversized_docx_zip() -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")
        archive.writestr("word/document.xml", "x" * 2048)
    return output.getvalue()


def _docx_bytes(text: str) -> bytes:
    """生成 docx2txt 可解析的最小 OOXML 文档。"""
    output = io.BytesIO()
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body><w:p><w:r><w:t>{escape(text)}</w:t></w:r></w:p></w:body>"
        "</w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    relationships = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/>'
        "</Relationships>"
    )
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", relationships)
        archive.writestr("word/document.xml", document)
    return output.getvalue()


async def test_real_pdf_docx_xlsx_and_utf8_txt_are_parsed_and_indexed(
    client, vectorstore, worker_once
):
    cases = (
        ("real.pdf", _text_pdf_bytes("PDF annual leave policy"), "application/pdf", "annual leave"),
        (
            "real.docx",
            _docx_bytes("DOCX expense policy"),
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "expense policy",
        ),
        (
            "real.xlsx",
            _xlsx_bytes(value="XLSX travel policy"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "travel policy",
        ),
        ("real.txt", "UTF-8 年假制度".encode("utf-8"), "text/plain", "年假制度"),
    )

    for filename, payload, mime, expected_text in cases:
        uploaded = await client.post(
            "/documents/upload",
            files={"file": (filename, payload, mime)},
        )
        assert uploaded.status_code == 201, uploaded.text
        assert await worker_once() is True
        doc_id = uploaded.json()["id"]
        detail = await client.get(f"/documents/{doc_id}")
        assert detail.json()["status"] == "indexed"
        stored_docs = [
            doc for doc in vectorstore.docstore._dict.values()
            if doc.metadata.get("doc_id") == doc_id
        ]
        assert stored_docs
        assert any(expected_text in doc.page_content for doc in stored_docs)
        assert all(doc.metadata["source"] == filename for doc in stored_docs)


async def test_upload_success_then_indexed(client, vectorstore, worker_once):
    """上传合法 txt：返回 201；后台索引跑完后状态变 indexed 且切片数 > 0。"""
    files = {"file": ("guide.txt", "公司报销流程：先在系统提单，再上传发票。".encode(), "text/plain")}
    resp = await client.post("/documents/upload", files=files)

    assert resp.status_code == 201
    body = resp.json()
    doc_id = body["id"]
    # POST 返回时状态还是 uploading（后台任务在响应后回填）
    assert body["status"] == "uploading"

    # 独立 Worker 领取持久化任务后，文档才进入 indexed。
    assert await worker_once() is True
    detail = await client.get(f"/documents/{doc_id}")
    assert detail.status_code == 200
    data = detail.json()
    assert data["status"] == "indexed"
    assert data["chunk_count"] > 0
    # 切片确实进了向量库
    assert len(vectorstore.docstore._dict) > 0
    first_id = list(vectorstore.docstore._dict.keys())[0]
    first_doc = vectorstore.docstore._dict[first_id]
    assert first_id == first_doc.metadata["chunk_id"]
    assert first_id.startswith(f"tenant-a:{doc_id}:0:")
    assert len(first_id.rsplit(":", 1)[1]) == 64


@pytest.mark.parametrize("filename", ["bad.zip", "evil.exe", "noext"])
async def test_upload_rejects_invalid_type(client, filename):
    """非 pdf/docx/xlsx/txt 一律 400，不落盘不索引。"""
    files = {"file": (filename, b"whatever", "application/octet-stream")}
    resp = await client.post("/documents/upload", files=files)
    assert resp.status_code == 400


async def test_upload_writes_db_and_lists(client, vectorstore):
    """上传后能在列表与详情接口查到该文档（验证写库）。"""
    files = {"file": ("policy.txt", b"hello knowledge base", "text/plain")}
    resp = await client.post("/documents/upload", files=files)
    doc_id = resp.json()["id"]

    listing = await client.get("/documents")
    assert listing.status_code == 200
    ids = [d["id"] for d in listing.json()]
    assert doc_id in ids

    detail = await client.get(f"/documents/{doc_id}")
    assert detail.json()["filename"] == "policy.txt"


async def test_get_missing_document_404(client):
    """查询不存在的文档返回 404。"""
    resp = await client.get("/documents/999999")
    assert resp.status_code == 404


async def test_300_character_filename_rejected(client):
    filename = f"{'a' * 296}.txt"
    response = await client.post(
        "/documents/upload",
        files={"file": (filename, b"safe", "text/plain")},
    )
    assert response.status_code == 400
    assert "文件名" in response.json()["detail"]


async def test_empty_file_rejected(client):
    response = await client.post(
        "/documents/upload",
        files={"file": ("empty.txt", b"", "text/plain")},
    )
    assert response.status_code == 400
    assert response.json()["detail"] == "空文件"


async def test_configured_file_size_limit_returns_413(client, monkeypatch):
    assert settings.max_file_size_bytes == 50 * 1024 * 1024
    monkeypatch.setattr(settings, "max_file_size_bytes", 10)
    response = await client.post(
        "/documents/upload",
        files={"file": ("large.txt", b"x" * 11, "text/plain")},
    )
    assert response.status_code == 413


async def test_extension_spoofing_rejected(client):
    response = await client.post(
        "/documents/upload",
        files={"file": ("fake.pdf", b"this is not a pdf", "application/pdf")},
    )
    assert response.status_code == 400
    assert "PDF" in response.json()["detail"]


async def test_mime_mismatch_rejected(client):
    response = await client.post(
        "/documents/upload",
        files={"file": ("fake.txt", b"plain text", "application/pdf")},
    )
    assert response.status_code == 400
    assert "MIME" in response.json()["detail"]


async def test_total_text_limit_returns_413(client, monkeypatch):
    monkeypatch.setattr(settings, "max_parsed_chars", 3)
    response = await client.post(
        "/documents/upload",
        files={"file": ("text.txt", b"four", "text/plain")},
    )
    assert response.status_code == 413
    assert "字符数" in response.json()["detail"]


async def test_pdf_page_limit_returns_413(client, monkeypatch):
    monkeypatch.setattr(settings, "max_pdf_pages", 1)
    response = await client.post(
        "/documents/upload",
        files={"file": ("pages.pdf", _pdf_bytes(2), "application/pdf")},
    )
    assert response.status_code == 413
    assert "页上限" in response.json()["detail"]


async def test_xlsx_cell_and_sheet_limits(client, monkeypatch):
    monkeypatch.setattr(settings, "max_xlsx_cells", 1)
    cells = await client.post(
        "/documents/upload",
        files={
            "file": (
                "cells.xlsx",
                _xlsx_bytes(cells=2),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert cells.status_code == 413
    assert "单元格" in cells.json()["detail"]

    monkeypatch.setattr(settings, "max_xlsx_cells", 100)
    monkeypatch.setattr(settings, "max_xlsx_sheets", 1)
    sheets = await client.post(
        "/documents/upload",
        files={
            "file": (
                "sheets.xlsx",
                _xlsx_bytes(sheets=2),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert sheets.status_code == 413
    assert "工作表" in sheets.json()["detail"]


async def test_archive_expansion_limit_returns_413(client, monkeypatch):
    monkeypatch.setattr(settings, "max_archive_uncompressed_bytes", 100)
    monkeypatch.setattr(settings, "max_archive_compression_ratio", 10_000.0)
    response = await client.post(
        "/documents/upload",
        files={
            "file": (
                "bomb.docx",
                _oversized_docx_zip(),
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
        },
    )
    assert response.status_code == 413
    assert "展开后" in response.json()["detail"]


async def test_successful_file_promoted_out_of_quarantine(
    client, vectorstore, worker_once
):
    response = await client.post(
        "/documents/upload",
        files={"file": ("trusted.txt", b"trusted text", "text/plain")},
    )
    assert response.status_code == 201
    assert await worker_once() is True
    async with database_module.AsyncSessionLocal() as db:
        document = (
            await db.execute(select(Document).where(Document.id == response.json()["id"]))
        ).scalar_one()
    parts = Path(document.file_path).parts
    assert "documents" in parts
    assert "quarantine" not in parts


async def test_parser_error_is_sanitized(tmp_path):
    missing = tmp_path / "private" / "secret.txt"
    result = await ingest_document(
        doc_id=1,
        file_path=str(missing),
        source="secret.txt",
        tenant_id="tenant-a",
        uploaded_by="user-a",
    )
    assert result.success is False
    assert result.error_msg == "文档解析失败"
    assert str(tmp_path) not in result.error_msg


def test_parser_uses_independent_process_pool():
    from concurrent.futures import ProcessPoolExecutor

    assert isinstance(get_parser_pool(), ProcessPoolExecutor)


async def test_parser_timeout_is_explicit_and_stays_quarantined(
    client,
    monkeypatch,
    worker_once,
):
    monkeypatch.setattr(settings, "ingest_job_max_attempts", 1)
    monkeypatch.setattr(settings, "parser_timeout_seconds", 0.000001)
    response = await client.post(
        "/documents/upload",
        files={"file": ("timeout.txt", b"safe text", "text/plain")},
    )
    assert response.status_code == 201
    assert await worker_once() is True
    detail = await client.get(f"/documents/{response.json()['id']}")
    assert detail.json()["status"] == "failed"
    assert detail.json()["error_msg"] == "文档解析超时"
    async with database_module.AsyncSessionLocal() as db:
        document = await db.get(Document, response.json()["id"])
    assert "quarantine" in Path(document.file_path).parts


async def test_malware_scanner_can_reject_quarantined_file(client):
    class RejectingScanner:
        async def scan(self, path: Path) -> MalwareScanResult:
            return MalwareScanResult(clean=False, scanned=True)

    configure_malware_scanner(RejectingScanner())
    try:
        response = await client.post(
            "/documents/upload",
            files={"file": ("malware.txt", b"test payload", "text/plain")},
        )
    finally:
        reset_malware_scanner()
    assert response.status_code == 400
    assert "恶意软件" in response.json()["detail"]


async def test_slow_upload_background_work_does_not_block_health(
    client,
    monkeypatch,
    worker_once,
):
    started = asyncio.Event()
    release = asyncio.Event()

    async def slow_ingest(**kwargs) -> IngestResult:
        started.set()
        await release.wait()
        return IngestResult(success=True, chunk_count=1)

    monkeypatch.setattr("app.services.ingest_jobs.ingest_document", slow_ingest)
    upload = await client.post(
        "/documents/upload",
        files={"file": ("slow.txt", b"safe", "text/plain")},
    )
    assert upload.status_code == 201
    worker = asyncio.create_task(worker_once())
    await asyncio.wait_for(started.wait(), timeout=2)
    health = await asyncio.wait_for(client.get("/health"), timeout=0.5)
    assert health.status_code == 200
    release.set()
    assert await worker is True
